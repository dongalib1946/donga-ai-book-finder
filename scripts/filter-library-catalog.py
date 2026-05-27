import argparse
import json
import re
import unicodedata
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


def text(value):
    if value is None:
        return ""
    value = str(value).strip()
    if value.endswith(".0") and value[:-2].isdigit():
        return value[:-2]
    return value


def norm_basic(value):
    value = unicodedata.normalize("NFKC", text(value)).lower()
    return re.sub(r"\s+", " ", value).strip()


def norm_compact(value):
    value = unicodedata.normalize("NFKC", text(value)).lower()
    return "".join(ch for ch in value if ch.isalnum())


def norm_author(value):
    value = norm_basic(value)
    return re.sub(r"\s*[,;]+$", "", value).strip()


def year4(value):
    match = re.search(r"(?:19|20)\d{2}", text(value))
    return match.group(0) if match else norm_basic(value)


def call_base(value):
    value = norm_basic(value)
    value = re.sub(r"\s*=\s*\d+\s*$", "", value)
    value = re.sub(r"\s+c\.?\s*\d+\s*$", "", value)
    value = re.sub(r"\s+copy\s*\d+\s*$", "", value)
    return value.strip()


def book_key(row):
    title = norm_compact(row["title"])
    author = norm_author(row["author"])
    year = year4(row["year"])
    call = call_base(row["callNo"])
    if call:
        return ("title-author-year-callbase", title, author, year, call)
    return ("title-author-year", title, author, year)


def has_general_reading_signal(title):
    return bool(re.search(
        r"교양|에세이|소설|시집|문학|세계사|역사|철학|심리|이야기|비밀|여행|인문|사람|삶|세상|사회|문화|예술|"
        r"재미있게|소설처럼|쉽게|알기 쉬운|최소한|처음|읽는 법|가이드북|회고록|인터뷰|산문",
        text(title),
        re.I,
    ))


def exclusion_reason(row):
    title = text(row["title"])
    compact_title = norm_compact(title)
    author = text(row["author"])
    call = text(row["callNo"])
    year = text(row["year"])
    full = " ".join([title, author, call, year])

    if re.search(r"웹툰|그래픽\s*노블|그래픽노블|코믹스|만화로\s*보는|도올\s*만화|만화논어|만화맹자|\bcomic(?:s)?\b|\bmanga\b|\bcartoon\b", full, re.I):
        return "comic"

    if re.search(
        r"수험서|수험\s*대비|문제집|기출|모의고사|실전모의|봉투모의|내신|대입|검정고시|국가고시|"
        r"의사국가고시|간호사국가고시|약사국가고시|치과의사국가고시|공무원\s*(?:국어|영어|한국사|행정법|행정학|헌법|사회|과학|수학|기출|문제|모의)|"
        r"공인중개사|임용고시|경찰공무원|소방공무원|인적성|"
        r"한능검|한국사능력검정|컴활|정보처리기사|자격증|기사\s*(?:필기|실기)|세무사|회계사|노무사|감정평가사|변리사|법무사|행정사",
        title,
        re.I,
    ):
        return "exam_prep"
    if re.search(r"(?<![가-힣])(?:수능|토익|토플|텝스)(?![가-힣])", title):
        return "exam_prep"
    if re.search(r"(?<![A-Za-z])(?:PSAT|LEET|MEET|DEET|NCS|TOEIC|TOEFL|IELTS|TEPS|JLPT|HSK)(?![A-Za-z])", title):
        return "exam_prep"

    low_demand_patterns = [
        r"경기\s*민속지|민속지|민속\s*조사|민속조사보고서|민속\s*자료|(?:^|[\s(:])(?:마을지|읍지|군지|시지|향토지)(?:$|[\s):])",
        r"조사\s*보고서|조사보고서|발굴\s*조사|발굴조사|시굴\s*조사|시굴조사|지표\s*조사|지표조사|정밀\s*발굴|문화재\s*조사|문화유적|매장문화재|유적\s*조사",
        r"(?:[가-힣]{2,}(?:리|동|면|읍|군|시)\s*)?유적\b|유물산포지|분묘군|고분군|패총|성곽\s*조사|사지\s*조사",
        r"연구\s*보고서|연구보고서|결과\s*보고서|결과보고서|최종\s*보고서|중간\s*보고서|사업\s*보고서|실태\s*조사|현황\s*조사",
        r"통계\s*연보|통계연보|연보\b|자료집\b|회의록\b|학술대회|심포지엄|워크숍|세미나\s*자료|공청회|토론회\s*자료",
    ]
    if any(re.search(pattern, full, re.I) for pattern in low_demand_patterns):
        if not has_general_reading_signal(title):
            return "low_demand_report_local"

    academic_field = re.search(
        r"간호|의학|약학|치의학|보건|임상|해부|생리|병리|혈액투석|유기화학|무기화학|물리화학|생화학|화학|물리학|수학|통계|"
        r"확률|미적분|선형대수|재료학|기계|전기|전자|회로|공학|토목|건축|컴퓨터|프로그래밍|데이터베이스|알고리즘|법학|"
        r"법이론|법철학|행정학|경영학|회계|세무|경제학|교육학|사회복지|체육학|운동영양|연구방법",
        full,
        re.I,
    )
    textbook_cue = re.search(
        r"개론|원론|각론|총론|입문|기초|기본|실습|실무|연습|워크북|핸드북|매뉴얼|연구\s*방법|방법론|"
        r"이론과\s*실제|진단|중재|가이드|최신|신편|이공학도|전공자|학습|강의|수업|제\s*\d+\s*판|=",
        title,
        re.I,
    )
    strong_textbook = re.search(
        r"간호학|간호진단|혈액투석간호|여성건강간호|유기화학|무기화학|물리화학|생화학|기계재료학|전기기기|회로이론|"
        r"확률\s*및\s*통계|미적분|선형대수|해부학|생리학|병리학|약리학|여행지리",
        title,
        re.I,
    )
    if strong_textbook:
        return "academic_textbook"
    if academic_field and textbook_cue and not has_general_reading_signal(title):
        return "academic_textbook"
    if re.search(r"(?:^|[\s(])(?:전공서|전공책|전공\s*교재|학부교재|대학강의|강의교재)(?:$|[\s):])", title, re.I):
        return "academic_textbook"
    if re.search(r"전공서|전공책|전공교재|학부교재|대학강의|강의교재", compact_title, re.I):
        return "academic_textbook"

    return ""


def load_excel_rows(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = []
    for raw in sheet.iter_rows(min_row=5, values_only=True):
        reg_no, title, author, call_no, year, location, shelved_at, _stop = (list(raw) + [""] * 8)[:8]
        if not text(title):
            continue
        rows.append({
            "regNo": text(reg_no),
            "title": text(title),
            "author": text(author),
            "callNo": text(call_no),
            "year": text(year),
            "location": text(location),
            "shelvedAt": text(shelved_at),
        })
    return rows


def load_pool_titles(path):
    with path.open("r", encoding="utf-8") as file:
        data = json.load(file)
    entries = data if isinstance(data, list) else data.get("entries", [])
    return {norm_compact(entry.get("title") or "") for entry in entries if norm_compact(entry.get("title") or "")}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--pool", default="netlify/data/library-pool.json")
    parser.add_argument("--output")
    args = parser.parse_args()

    rows = load_excel_rows(Path(args.excel))

    unique = {}
    for row in rows:
        unique.setdefault(book_key(row), row)
    deduped = list(unique.values())

    pool_titles = load_pool_titles(Path(args.pool))
    after_pool = [row for row in deduped if norm_compact(row["title"]) not in pool_titles]

    reason_counts = Counter()
    examples = {}
    kept = []
    for row in after_pool:
        reason = exclusion_reason(row)
        if reason:
            reason_counts[reason] += 1
            examples.setdefault(reason, []).append(f"{row['title']} / {row['author']} / {row['year']}")
        else:
            kept.append(row)

    if args.output:
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with output_path.open("w", encoding="utf-8") as file:
            json.dump(kept, file, ensure_ascii=False, indent=2)

    print(json.dumps({
        "sourceExcel": str(Path(args.excel)),
        "physicalRows": len(rows),
        "afterCopyDedup": len(deduped),
        "removedByLibraryPoolTitleOnly": len(deduped) - len(after_pool),
        "afterLibraryPoolRemoval": len(after_pool),
        "excludedByRulesTotal": sum(reason_counts.values()),
        "excludedByReason": dict(reason_counts),
        "finalRemaining": len(kept),
        "output": args.output or "",
        "examplesByReason": {key: values[:8] for key, values in examples.items()},
        "sampleRemaining": [f"{row['title']} / {row['author']} / {row['year']}" for row in kept[:12]],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
